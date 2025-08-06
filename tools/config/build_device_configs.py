#!/usr/bin/env python3
"""
Build device mappings and XML templates from BMC Server List Excel file
"""

import openpyxl
import yaml
import os
from pathlib import Path

def read_server_data():
    """Read server data from Excel file"""
    try:
        wb = openpyxl.load_workbook('BMC Server List.xlsx')
        sheet = wb['Server Types']
        
        # Get headers
        headers = []
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=col).value
            headers.append(value if value else f'Column_{col}')
        
        # Read data rows
        servers = []
        for row_num in range(2, sheet.max_row + 1):
            server = {}
            for col_num, header in enumerate(headers, 1):
                value = sheet.cell(row=row_num, column=col_num).value
                if value is not None:
                    server[header] = value
            if server.get('_id'):  # Only include rows with valid IDs
                servers.append(server)
        
        print(f"✅ Loaded {len(servers)} server configurations")
        return servers
        
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return []

def categorize_server_by_specs(server):
    """Categorize server into small/medium/large based on specs"""
    cores = server.get('total cores', 0)
    ram_gb = server.get('ramInGb', 0)
    
    if cores >= 16 or ram_gb >= 128:
        return 'large'
    elif cores >= 8 or ram_gb >= 64:
        return 'medium'
    else:
        return 'small'

def get_motherboard_from_cpu(cpu_name, server_vendor):
    """Estimate motherboard based on CPU and vendor"""
    if not cpu_name or server_vendor != 'supermicro':
        return ['Unknown']
    
    cpu_name = str(cpu_name).upper()
    
    # Intel Xeon E-series mapping
    if 'E-22' in cpu_name:  # Coffee Lake
        return ['X11SCE-F', 'X11SCM-F']
    elif 'E-23' in cpu_name:  # Comet Lake
        return ['X12STE-F', 'X12SCE-F']
    elif 'XEON' in cpu_name and 'SILVER' in cpu_name:
        return ['X12DPT-B', 'X12DPG-QT']
    elif 'XEON' in cpu_name and 'GOLD' in cpu_name:
        return ['X12DPT-B', 'X13DET-B']
    else:
        return ['X11DPT-B', 'X11DPFR-SN']

def determine_power_profile(server_id, cpu_freq, cores):
    """Determine power profile based on server characteristics"""
    if 'gaming' in server_id.lower() or cpu_freq >= 4.0:
        return 'performance'
    elif cores <= 4:
        return 'balanced'
    else:
        return 'performance'

def get_boot_mode(server_vendor, cpu_name):
    """Determine boot mode based on hardware"""
    # Modern servers should use UEFI
    if server_vendor == 'supermicro':
        return 'uefi'
    return 'legacy'

def build_device_mappings(servers):
    """Build device mappings YAML from server data"""
    device_types = {}
    
    for server in servers:
        server_id = server.get('_id', '')
        if not server_id:
            continue
            
        # Extract specs
        cores = server.get('total cores', 0)
        ram_gb = server.get('ramInGb', 0)
        cpu_name = server.get('cpuName', '')
        cpu_freq = server.get('cpuFrequency', 0)
        server_vendor = server.get('server_vendor', 'unknown')
        storage = server.get('storage', '')
        sgx_tdx = server.get('SGX/TDX enabled', '')
        
        # Build description
        size_category = categorize_server_by_specs(server)
        description = f"{size_category.title()} server - {cores} cores, {ram_gb}GB RAM"
        if cpu_name:
            description += f", {cpu_name}"
        
        # Build configuration
        device_config = {
            'description': description,
            'motherboards': get_motherboard_from_cpu(cpu_name, server_vendor),
            'boot_configs': {
                'boot_mode': get_boot_mode(server_vendor, cpu_name),
                'pxe_boot': True,
                'secure_boot': False
            },
            'cpu_configs': {
                'hyperthreading': True,
                'power_profile': determine_power_profile(server_id, cpu_freq, cores),
                'turbo_boost': True
            },
            'memory_configs': {
                'ecc_enabled': True,
                'memory_speed': 'auto'
            },
            'hardware_specs': {
                'cpu_cores': cores,
                'ram_gb': ram_gb,
                'cpu_name': cpu_name,
                'cpu_frequency': cpu_freq,
                'storage': storage,
                'network': server.get('network', ''),
                'vendor': server_vendor
            }
        }
        
        # Add security features if available
        if sgx_tdx:
            device_config['security_configs'] = {
                'sgx_enabled': 'SGX' in sgx_tdx.upper(),
                'tdx_enabled': 'TDX' in sgx_tdx.upper()
            }
        
        device_types[server_id] = device_config
    
    return {'device_types': device_types}

def generate_xml_template(server_id, config):
    """Generate XML template for BIOS configuration"""
    specs = config.get('hardware_specs', {})
    boot_configs = config.get('boot_configs', {})
    cpu_configs = config.get('cpu_configs', {})
    memory_configs = config.get('memory_configs', {})
    security_configs = config.get('security_configs', {})
    
    xml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<!-- Auto-generated BIOS template for {server_id} -->
<!-- {config.get('description', 'No description')} -->
<BiosCfg>
    <Menu name="Advanced">
        <!-- CPU Configuration -->
        <Setting name="Intel Hyper-Threading Technology" selectedOption="{'Enabled' if cpu_configs.get('hyperthreading') else 'Disabled'}" />
        <Setting name="Intel Turbo Boost Technology" selectedOption="{'Enabled' if cpu_configs.get('turbo_boost') else 'Disabled'}" />
        <Setting name="Power Management" selectedOption="{cpu_configs.get('power_profile', 'balanced').title()}" />
        
        <!-- Memory Configuration -->
        <Setting name="Memory RAS Configuration">
            <Setting name="Memory Correctable Error" selectedOption="{'Enabled' if memory_configs.get('ecc_enabled') else 'Disabled'}" />
        </Setting>
        
        <!-- Boot Configuration -->
        <Setting name="Boot mode select" selectedOption="{'UEFI' if boot_configs.get('boot_mode') == 'uefi' else 'Legacy'}" />
        <Setting name="PXE Boot" selectedOption="{'Enabled' if boot_configs.get('pxe_boot') else 'Disabled'}" />
        <Setting name="Secure Boot" selectedOption="{'Enabled' if boot_configs.get('secure_boot') else 'Disabled'}" />
"""
    
    # Add security configurations if available
    if security_configs:
        xml_content += """        
        <!-- Security Configuration -->"""
        if security_configs.get('sgx_enabled'):
            xml_content += """
        <Setting name="Intel Software Guard Extensions (SGX)" selectedOption="Enabled" />"""
        if security_configs.get('tdx_enabled'):
            xml_content += """
        <Setting name="Intel Trust Domain Extensions (TDX)" selectedOption="Enabled" />"""
    
    # Add performance optimizations based on server category
    cores = specs.get('cpu_cores', 0)
    if cores >= 16:
        xml_content += """
        
        <!-- High-Performance Optimizations -->
        <Setting name="C-State" selectedOption="Disabled" />
        <Setting name="Package C State" selectedOption="C0/C1 state" />
        <Setting name="CPU P State Control" selectedOption="Disabled" />"""
    
    xml_content += """
    </Menu>
    
    <!-- IPMI Configuration -->
    <Menu name="IPMI">
        <Setting name="BMC Network Configuration" selectedOption="Static" />
        <Setting name="IPMI LAN Selection" selectedOption="Dedicated" />
    </Menu>
</BiosCfg>"""
    
    return xml_content

def main():
    print("🚀 Building device mappings and XML templates from BMC Server List...")
    
    # Read server data
    servers = read_server_data()
    if not servers:
        print("❌ No server data found")
        return
    
    # Build device mappings
    print("\n📋 Building device mappings...")
    device_mappings = build_device_mappings(servers)
    
    # Save device mappings
    mappings_file = 'configs/bios/device_mappings_generated.yaml'
    os.makedirs('configs/bios', exist_ok=True)
    
    with open(mappings_file, 'w') as f:
        yaml.dump(device_mappings, f, default_flow_style=False, indent=2)
    
    print(f"✅ Saved device mappings to {mappings_file}")
    print(f"   Generated {len(device_mappings['device_types'])} device types")
    
    # Generate XML templates
    print("\n🛠️  Generating XML templates...")
    xml_dir = 'configs/bios/xml_templates_generated'
    os.makedirs(xml_dir, exist_ok=True)
    
    generated_count = 0
    for server_id, config in device_mappings['device_types'].items():
        xml_content = generate_xml_template(server_id, config)
        xml_file = f"{xml_dir}/{server_id}.xml"
        
        with open(xml_file, 'w') as f:
            f.write(xml_content)
        
        generated_count += 1
    
    print(f"✅ Generated {generated_count} XML templates in {xml_dir}/")
    
    # Show sample configurations
    print("\n📊 Sample Device Types Generated:")
    sample_types = list(device_mappings['device_types'].keys())[:5]
    for device_type in sample_types:
        config = device_mappings['device_types'][device_type]
        specs = config.get('hardware_specs', {})
        print(f"  • {device_type}: {specs.get('cpu_cores', 0)} cores, {specs.get('ram_gb', 0)}GB RAM")
    
    print(f"\n🎉 Complete! Generated configurations for {len(device_mappings['device_types'])} server types")
    print("\nNext steps:")
    print("1. Review the generated files")
    print("2. Merge configs/bios/device_mappings_generated.yaml with your existing device_mappings.yaml")
    print("3. Copy XML templates from xml_templates_generated/ to xml_templates/")
    print("4. Test BIOS configurations with your hardware")

if __name__ == "__main__":
    main()
