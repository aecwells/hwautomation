#!/usr/bin/env python3
"""
Simple Excel reader for BMC Server List
"""

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

def read_bmc_server_list():
    try:
        # Load the workbook
        wb = openpyxl.load_workbook('BMC Server List.xlsx')
        print('✅ Excel file loaded successfully!')
        print(f'Sheet names: {wb.sheetnames}')
        
        # Get the active sheet
        sheet = wb.active
        print(f'Active sheet: {sheet.title}')
        print(f'Dimensions: {sheet.max_row} rows x {sheet.max_column} columns')
        print()
        
        # Read all data
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        
        if data:
            headers = data[0]
            print('=== Headers ===')
            for i, header in enumerate(headers, 1):
                print(f'{i}: {header}')
            print()
            
            print('=== Sample Data (First 3 rows) ===')
            for i, row in enumerate(data[1:4], 1):  # Skip header, show first 3 data rows
                print(f'Row {i}:')
                for j, (header, value) in enumerate(zip(headers, row)):
                    if value is not None and str(value).strip():
                        print(f'  {header}: {value}')
                print()
                
            print(f'Total data rows: {len(data) - 1}')
        
        return data
        
    except Exception as e:
        print(f'❌ Error reading Excel file: {e}')
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    read_bmc_server_list()
